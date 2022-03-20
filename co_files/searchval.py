from openpyxl.workbook import Workbook
from openpyxl import load_workbook

van=[1001,200026110,'IHT3026','IHT3099']
mylist = [str(int) for int in van]

wb=load_workbook('ABC1001.xlsx')
ws=wb.active
L=(f'{ws["A2"].value}')

def Search(mylist):    
    for i in range(0, n):  
        if (mylist[i] == L):
            print("Matched Available value")

n=len(mylist)
Search(mylist)
